"""In-memory builders for synthetic Standard Format workbooks (no PII)."""

from __future__ import annotations

import io
from datetime import datetime

import openpyxl

# Deliberately stale labels (as in the real workbook): wrong years, and row 2
# even carries a wrong month. Parsers must map by row position instead.
STALE_LABELS = [
    "03/2023", "10/2022", "05/2023", "06/2023", "07/2023", "08/2023",
    "09/2023", "10/2023", "11/2023", "12/2023", "01/2024", "02/2024",
]
CORRECT_LABELS_2023 = [
    "03/2022", "04/2022", "05/2022", "06/2022", "07/2022", "08/2022",
    "09/2022", "10/2022", "11/2022", "12/2022", "01/2023", "02/2023",
]

_HEADER = ["", "Salaris", "Leave pay", "Oortyd", "Bonus", "Reistoelaag",
           "Verlof", "Bruto salaris", "PAYE", "SDL", "UIF", "UIF"]


def _month_row(label, salaris=0.0, leave=0.0, oortyd=0.0, bonus=0.0,
               reis=0.0, verlof=0.0, bruto=None, uif1=None):
    total = salaris + leave + oortyd + bonus + reis + verlof
    if bruto is None:
        bruto = total
    if uif1 is None:
        uif1 = round(min(total, 17712.0) * 0.01, 4)
    return [label, salaris, leave, oortyd, bonus, reis, verlof,
            bruto, 0, round(bruto * 0.01, 4), uif1, uif1 * 2]


def _block(ws, nr, name, surname, id_no, start, end, reason, month_rows):
    ws.append(["Employee Nr.:", nr])
    ws.append(["Name:", name])
    ws.append(["Surname:", surname])
    ws.append(["ID No.:", id_no])
    ws.append(["Income Tax Nr.:", "1234567890"])
    ws.append(["Start date:", start])
    ws.append(["End date:", end])
    ws.append(["Reason:", reason])
    ws.append([])
    ws.append(["", "", "", "", "", "", "", "", "", "1%", "1%", "2%"])
    ws.append(_HEADER)
    ws.append([])
    for row in month_rows:
        ws.append(row)
    ws.append([""] + [sum(r[i] for r in month_rows) for i in range(1, 12)])
    ws.append([])


def build_payroll_workbook() -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, labels in (("2023", CORRECT_LABELS_2023), ("2025", STALE_LABELS)):
        ws = wb.create_sheet(sheet_name)
        ws.append(["Company:", "Acme Electrical (Pty) Ltd"])
        ws.append(["Financial year:", 2024])          # stale on purpose
        ws.append(["PAYE No.:", "7012345678"])
        ws.append(["UIF No.:", "0123456/7"])
        ws.append([])

        if sheet_name == "2023":
            emp1_rows = [_month_row(labels[i]) for i in range(11)]
            emp1_rows.append(_month_row(labels[11], salaris=6000.0))          # February
        else:
            emp1_rows = [
                _month_row(labels[0], salaris=10000.0),                       # March: clean
                _month_row(labels[1], salaris=8300.0, uif1=75.00),           # April: UIF divergence
                _month_row(labels[2], salaris=5000.0, bruto=5100.0),          # May: bruto mismatch
                _month_row(labels[3], salaris=20000.0),                       # June: capped, no warning
                *(_month_row(labels[i]) for i in range(4, 9)),
                _month_row(labels[9], salaris=5500.0, leave=5200.0, bonus=4600.0),  # December
                _month_row(labels[10]),
                _month_row(labels[11]),
            ]
        _block(ws, "001", "Petrus Johannes", "Botha", "AB123456",
               "03/09/2018", "N/A", "-", emp1_rows)
        _block(ws, "002", "Piet", "van Wyk", "9103105023081",
               datetime(2018, 4, 3), "05/04/2022", "Death",
               [_month_row(labels[i]) for i in range(12)])

    wb.create_sheet("Notes").append(["scratch", "not a year sheet"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_payroll_workbook_missing_month() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2024"
    _block(ws, "001", "Jan", "Marais", "9001015009087",
           "01/02/2020", "N/A", "-",
           [_month_row(f"{m:02d}/2023") for m in range(1, 12)])   # 11 rows
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_master_workbook() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee details"
    ws.append(["Employee Number", "Name", "Surname", "ID No. / Passport Number",
               "Date of Birth", "Income Tax No.", "Start Date", "End Date", "Reason"])
    ws.append(["001", "Petrus Johannes", "Botha", "AB123456",
               "14/02/1979", "1234567801", "03/09/2018", "N/A", "-"])
    ws.append(["002", "Daniel Sipho ", "Nkosi", 9103105023081,
               datetime(1991, 3, 10), "1122334455", datetime(2018, 4, 3), "N/A", "-"])
    ws.append(["003", "Andries ", "Fourie", "8807235112083",
               "23/07/1988", "2233445566", "02/03/2015", "07/11/2022", "Death"])
    ws.append(["009", "Willem Karel ", "Van Der Walt", "6512315678087",
               "31/12/1965", "3344556677", "01/07/2019", "N/A", "-"])
    ws.append(["010"])          # number-only stub row — must be skipped
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_payroll_workbook_trailing_empty_block() -> bytes:
    """One valid block, then a block whose Employee Nr. cell is empty
    (stale template rows) — the scan must stop there."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2024"
    _block(ws, "001", "Jan", "Marais", "9001015009087",
           "01/02/2020", "N/A", "-",
           [_month_row(f"{m:02d}/2023", salaris=6000.0) for m in range(1, 13)])
    _block(ws, None, "", "", "", "", "N/A", "-",
           [_month_row(f"{m:02d}/2023") for m in range(1, 13)])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_payroll_workbook_duplicate_code() -> bytes:
    """Two full 12-month blocks both numbered 001 — a data error."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2024"
    for _ in range(2):
        _block(ws, "001", "Jan", "Marais", "9001015009087",
               "01/02/2020", "N/A", "-",
               [_month_row(f"{m:02d}/2023", salaris=6000.0) for m in range(1, 13)])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_master_workbook_duplicate_code() -> bytes:
    """Master sheet with two full identity rows both numbered 001."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee details"
    ws.append(["Employee Number", "Name", "Surname", "ID No. / Passport Number",
               "Date of Birth", "Income Tax No.", "Start Date", "End Date", "Reason"])
    ws.append(["001", "Jan", "Marais", "9001015009087",
               "01/01/1990", "1234567890", "01/02/2020", "N/A", "-"])
    ws.append(["001", "Piet", "van Wyk", "9103105023081",
               "10/03/1991", "1234567890", "03/04/2021", "N/A", "-"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
