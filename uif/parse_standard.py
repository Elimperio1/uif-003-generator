"""
Parsers for "Standard Format" xlsx workbooks — hand-kept payroll for clients
not on Sage (first case: the client).

Two workbooks:
  - payroll workbook: one sheet per tax year ("2022".."2027"), each holding
    per-employee identity blocks followed by a 12-row month table;
  - master workbook: an "Employee details" sheet, one row per employee.

Data-quality rulings (docs/superpowers/specs/2026-08-11-standard-format-design.md):
  - tax year comes from the sheet tab name and month from row position; the
    in-sheet month labels are stale template copies and are ignored;
  - UIF is recomputed from the earning columns; months where the sheet's own
    UIF column differs are reported as soft warnings.
"""

from __future__ import annotations

import datetime as _dt
import io
import re

import openpyxl

from .models import TAX_YEAR_MONTHS, UIF_REMUNERATION_CAP, EmployeeRecord, YtdRecord
from .parse_employees import extract_first_name, parse_employee_code, parse_id_number


def detect_format(file_bytes: bytes) -> str:
    """'standard' for xlsx (zip magic), 'sage' for anything else (CSV)."""
    return "standard" if file_bytes[:4] == b"PK\x03\x04" else "sage"


def _load(file_bytes: bytes):
    return openpyxl.load_workbook(
        io.BytesIO(file_bytes), data_only=True, read_only=True
    )


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def list_year_sheets(file_bytes: bytes) -> list[str]:
    """Sheet names that are 4-digit years, sorted ascending."""
    wb = _load(file_bytes)
    try:
        names = [n for n in wb.sheetnames if re.fullmatch(r"\d{4}", n.strip())]
    finally:
        wb.close()
    return sorted(names, key=int)


def tax_year_end_year(sheet_name: str) -> int:
    """Ruling: the sheet tab name IS the tax-year end year."""
    return int(str(sheet_name).strip())


def read_company_header(file_bytes: bytes, sheet_name: str) -> dict[str, str]:
    """Company / PAYE / UIF cells from the top of a payroll year sheet."""
    out = {"company": "", "paye": "", "uif": ""}
    labels = {"Company:": "company", "PAYE No.:": "paye", "UIF No.:": "uif"}
    wb = _load(file_bytes)
    try:
        for row in wb[sheet_name].iter_rows(max_row=8, values_only=True):
            key = labels.get(_text(row[0] if row else None))
            if key and len(row) > 1:
                out[key] = _text(row[1])
    finally:
        wb.close()
    return out


MASTER_SHEET = "Employee details"
_EMPTY_VALUES = {"", "n/a", "-"}

EARNING_COLUMNS = ("Salaris", "Leave pay", "Oortyd", "Bonus", "Reistoelaag", "Verlof")
_MONTH_LABEL_RE = re.compile(r"\d{2}/\d{4}")


def _num(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("\xa0", "").replace(" ", "").replace(",", "").strip()
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _date(value) -> str:
    """Cell to YYYYMMDD. Accepts Excel datetimes and 'DD/MM/YYYY' strings;
    'N/A', '-' and blank mean no date."""
    if isinstance(value, (_dt.datetime, _dt.date)):
        return f"{value:%Y%m%d}"
    text = _text(value)
    if text.lower() in _EMPTY_VALUES:
        return ""
    match = re.search(r"(\d{2})/(\d{2})/(\d{4})", text)
    return f"{match.group(3)}{match.group(2)}{match.group(1)}" if match else ""


def parse_employees(file_bytes: bytes) -> dict[str, EmployeeRecord]:
    """Parse the master workbook's 'Employee details' sheet."""
    wb = _load(file_bytes)
    try:
        if MASTER_SHEET not in wb.sheetnames:
            raise ValueError(
                f'Could not find the "{MASTER_SHEET}" sheet in the employee '
                f"master workbook."
            )
        rows = list(wb[MASTER_SHEET].iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        raise ValueError(f'The "{MASTER_SHEET}" sheet is empty.')

    header = [_text(c).lower() for c in rows[0]]

    def find_col(prefix: str) -> int | None:
        for i, cell in enumerate(header):
            if cell.startswith(prefix):
                return i
        return None

    cols = {
        "code": find_col("employee number"),
        "name": find_col("name"),
        "surname": find_col("surname"),
        "id": find_col("id no"),
        "dob": find_col("date of birth"),
        "start": find_col("start date"),
        "end": find_col("end date"),
    }
    missing = [k for k in ("code", "name", "surname", "id") if cols[k] is None]
    if missing:
        raise ValueError(
            f'The "{MASTER_SHEET}" sheet is missing expected columns: '
            f"{', '.join(sorted(missing))}."
        )

    def cell(row, key):
        i = cols[key]
        return row[i] if i is not None and i < len(row) else None

    records: dict[str, EmployeeRecord] = {}
    for row in rows[1:]:
        code = parse_employee_code(cell(row, "code"))
        first = extract_first_name(_text(cell(row, "name")))
        surname = _text(cell(row, "surname"))
        if not code or (not first and not surname):
            continue  # stub rows: an employee number with no identity yet
        if code in records:
            raise ValueError(
                f'Employee number {code} appears more than once on the '
                f'"{MASTER_SHEET}" sheet.'
            )

        raw_id = _text(cell(row, "id"))
        digits = parse_id_number(raw_id)
        if len(digits) == 13 and digits.isdigit():
            id_number, passport = digits, ""
        else:
            id_number = ""
            passport = "" if raw_id.lower() in _EMPTY_VALUES else raw_id

        end_date = _date(cell(row, "end"))
        records[code] = EmployeeRecord(
            employee_code=code,
            surname=surname,
            first_names=first,
            id_number=id_number,
            passport_number=passport,
            date_of_birth=_date(cell(row, "dob")),
            date_engaged=_date(cell(row, "start")),
            end_date=end_date,
            employee_status="Terminated" if end_date else "Normal",
            uif_status="Contributes",
        )
    return records


def parse_ytd(
    file_bytes: bytes, sheet_name: str
) -> tuple[dict[str, YtdRecord], list[str]]:
    """
    Parse one year sheet of the payroll workbook into
    ({employee_code: YtdRecord}, soft_warnings).

    Months map by row position (1st table row = March .. 12th = February);
    the MM/YYYY labels are stale template copies and only gate which rows
    count as month rows. A block without exactly 12 month rows is a
    blocking parse error.
    """
    wb = _load(file_bytes)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' was not found in the payroll workbook."
            )
        rows = list(wb[sheet_name].iter_rows(values_only=True))
    finally:
        wb.close()

    records: dict[str, YtdRecord] = {}
    warnings: list[str] = []

    current: YtdRecord | None = None
    name_part = surname_part = ""
    earn_cols: dict[int, str] = {}
    bruto_col: int | None = None
    uif1_col: int | None = None
    month_index = 0

    def finalise() -> None:
        if current is not None and month_index != 12:
            raise ValueError(
                f"Employee {current.employee_code} ({current.employee_name}) "
                f"on sheet '{sheet_name}' has {month_index} month rows, "
                f"expected 12."
            )

    for row in rows:
        first = _text(row[0] if row else None)

        if first == "Employee Nr.:":
            finalise()
            code = parse_employee_code(row[1] if len(row) > 1 else None)
            if not code:
                # Ruling: a block with an empty Employee Nr. cell ends the scan.
                current = None
                break
            if code in records:
                raise ValueError(
                    f"Employee code {code} appears more than once on "
                    f"sheet '{sheet_name}'."
                )
            current = YtdRecord(
                employee_code=code,
                employee_name="",
                status="Employed",
                earnings={m: {} for m in TAX_YEAR_MONTHS},
            )
            records[code] = current
            name_part = surname_part = ""
            earn_cols, bruto_col, uif1_col = {}, None, None
            month_index = 0
            continue
        if current is None:
            continue

        value = _text(row[1]) if len(row) > 1 else ""
        if first == "Name:":
            name_part = value
            current.employee_name = f"{name_part} {surname_part}".strip()
            continue
        if first == "Surname:":
            surname_part = value
            current.employee_name = f"{name_part} {surname_part}".strip()
            continue
        if first == "End date:":
            current.end_date = _date(row[1] if len(row) > 1 else None)
            if current.end_date:
                current.status = "Terminated"
            continue
        if first == "Reason:":
            current.reason = value
            if value.lower() == "death":
                warnings.append(
                    f"Employee {current.employee_code} "
                    f"({current.employee_name}): reason is 'Death' — "
                    f"pre-selected as 02 Deceased in Step 4; confirm before "
                    f"downloading."
                )
            continue

        cells = [_text(c) for c in row]
        if "Salaris" in cells and "Bruto salaris" in cells:
            earn_cols = {i: c for i, c in enumerate(cells) if c in EARNING_COLUMNS}
            bruto_col = cells.index("Bruto salaris")
            uif_positions = [i for i, c in enumerate(cells) if c == "UIF"]
            uif1_col = uif_positions[0] if uif_positions else None
            continue

        if earn_cols and _MONTH_LABEL_RE.fullmatch(first):
            month_index += 1
            if month_index > 12:
                continue  # finalise() reports the bad count
            month = TAX_YEAR_MONTHS[month_index - 1]

            row_total = 0.0
            for i, earning_name in earn_cols.items():
                amount = _num(row[i]) if i < len(row) else 0.0
                row_total += amount
                if amount:
                    current.earnings[month][earning_name] = (
                        current.earnings[month].get(earning_name, 0.0) + amount
                    )

            label = (
                f"Employee {current.employee_code} "
                f"({current.employee_name}), {month}"
            )
            if bruto_col is not None and bruto_col < len(row):
                bruto = _num(row[bruto_col])
                if abs(bruto - row_total) > 0.01:
                    warnings.append(
                        f"{label}: the sheet's Bruto salaris ({bruto:.2f}) "
                        f"does not equal the sum of the earning columns "
                        f"({row_total:.2f})."
                    )
            if uif1_col is not None and uif1_col < len(row):
                sheet_uif = _num(row[uif1_col])
                expected = min(row_total, UIF_REMUNERATION_CAP) * 0.01
                if abs(sheet_uif - expected) > 0.02:
                    warnings.append(
                        f"{label}: the sheet's UIF column ({sheet_uif:.2f}) "
                        f"is not 1% of the (capped) earnings "
                        f"({expected:.2f}) — the app uses the recomputed "
                        f"value."
                    )

    finalise()
    return records, warnings
